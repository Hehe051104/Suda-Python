import openpyxl as px
import random
import os
import sys

def get_resource_path(relative_path):
    """获取资源的绝对路径"""
    try:
        # PyInstaller创建临时文件夹,将路径存储在_MEIPASS中
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Helper function to safely convert a value to an integer.
# Handles None, non-numeric strings, etc., by returning 0.
def _safe_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return 0

class TxtFile:
    @classmethod
    def getGameInfo(cls):
        s=''
        try:
            f=open(get_resource_path('小测验游戏说明.txt'), encoding='gbk')
        except:
            return s
        else:
            while True:
                line=f.readline()
                if not line:
                    break
                s+=line+'\n'
            f.close()
            return s
    @classmethod
    def getMaxScore(cls):
        try:
            f=open(get_resource_path('Record.txt'), encoding='gbk')
        except:
            return -1
        else:
            r=f.readlines()
            f.close()
            return int(r[3][3:].strip())
    @classmethod
    def setNewRecord(cls,sno,sname,score):
        f=open(get_resource_path('Record.txt'),'w', encoding='gbk')
        f.write('最高得分记录如下: \n')
        f.write('学号: '+sno+'\n')
        f.write('姓名: '+sname+'\n')
        f.write('得分: '+str(score)+'\n')
        f.close()

class Stu:
    date=0
    sheet=0
    def __init__(self):
        Stu.date=px.load_workbook(get_resource_path('名单.xlsx'))
        Stu.sheet=Stu.date.active
    def getStu(self):
        stu_info=[]
        row_num=Stu.sheet.max_row
        rows=Stu.sheet['A2':'B%d'%row_num]
        for row in rows:
            sno_val = str(row[0].value).strip() if row[0].value is not None else ""
            sname_val = str(row[1].value).strip() if row[1].value is not None else ""
            stu_info.append((sno_val, sname_val))
        return stu_info
class QusAndAns:
    date=0
    sheet1=0
    sheet2=0
    sheet3=0
    sheet4=0
    cnt_select=0
    cnt_blank=0
    cnt_judge=0
    examName=''
    examTime=0
    totalSelect=0
    totalblank=0
    totalJudge=0
    iselectScore=0
    iselectblank=0
    ijudgeScore=0
    qus_select=[]
    ans_select=[]
    analyze_select=[]
    qus_blank = []
    ans_blank = []
    analyze_blank = []
    qus_judge = []
    ans_judge = []
    analyze_judge = []


    def __init__(self):
        QusAndAns.date=px.load_workbook(get_resource_path("Quiz.xlsx"))
        try: QusAndAns.sheet1=QusAndAns.date['选择题'] 
        except: QusAndAns.sheet1 = None
        try: QusAndAns.sheet2=QusAndAns.date['填空题']
        except: QusAndAns.sheet2 = None
        try: QusAndAns.sheet3=QusAndAns.date['测验信息']
        except: QusAndAns.sheet3 = None
        try: QusAndAns.sheet4=QusAndAns.date['判断题']
        except: QusAndAns.sheet4 = None

        if QusAndAns.sheet1: QusAndAns.cnt_select = _safe_int(QusAndAns.sheet1.cell(1, 8).value)
        if QusAndAns.sheet2: QusAndAns.cnt_blank = _safe_int(QusAndAns.sheet2.cell(1, 8).value)
        if QusAndAns.sheet4: QusAndAns.cnt_judge = _safe_int(QusAndAns.sheet4.cell(1, 8).value)

    # 获取测验名和测验时长
    def getEnameAndEtime(self):
        if not QusAndAns.sheet3: return ("未命名测验", 30)
        QusAndAns.examName=QusAndAns.sheet3.cell(1,2).value
        QusAndAns.examTime=_safe_int(QusAndAns.sheet3.cell(2,2).value)
        return (QusAndAns.examName,QusAndAns.examTime)

    # 获取选择题总分和每题分数
    def getTotalAndiScore(self):
        if not QusAndAns.sheet1: return (0, 0)
        QusAndAns.totalSelect=_safe_int(QusAndAns.sheet1.cell(1,4).value)
        QusAndAns.iselectScore=_safe_int(QusAndAns.sheet1.cell(1,6).value)
        return (QusAndAns.totalSelect,QusAndAns.iselectScore)

    # 获取填空题总分和每题分数
    def getTotalAndiScore1(self):
        if not QusAndAns.sheet2: return (0, 0)
        QusAndAns.totalblank = _safe_int(QusAndAns.sheet2.cell(1, 4).value)
        QusAndAns.iblankScore = _safe_int(QusAndAns.sheet2.cell(1, 6).value)
        return (QusAndAns.totalblank, QusAndAns.iblankScore)

    # 新增：获取判断题总分和每题分数
    def getTotalAndiScore2(self):
        if not QusAndAns.sheet4: return (0, 0)
        QusAndAns.totalJudge = _safe_int(QusAndAns.sheet4.cell(1, 4).value)
        QusAndAns.ijudgeScore = _safe_int(QusAndAns.sheet4.cell(1, 6).value)
        return (QusAndAns.totalJudge, QusAndAns.ijudgeScore)

    # 获取选择题
    def getQusOfSelect(self):
        if not QusAndAns.sheet1: return []
        if QusAndAns.qus_select: return QusAndAns.qus_select
        for row in range(3,3+QusAndAns.cnt_select):
            con=QusAndAns.sheet1['A%d'%row:'F%d'%row]
            temp1='、'+str(con[0][1].value)+'\n'
            temp2='A.'+str(con[0][2].value)+'\n'+'B. '+str(con[0][3].value)+'\n'+'C. '+str(con[0][4].value)+'\n'+'D. '+str(con[0][5].value)
            QusAndAns.qus_select.append((temp1+temp2))
        return QusAndAns.qus_select

    # 获取选择题答案
    def getAnsOfSelect(self):
        if not QusAndAns.sheet1: return []
        if QusAndAns.ans_select: return QusAndAns.ans_select
        for row in range(3,3+QusAndAns.cnt_select):
            con=QusAndAns.sheet1['G%d'%row:'G%d'%row]
            QusAndAns.ans_select.append(str(con[0][0].value))
        return QusAndAns.ans_select

    # 获取选择题解析
    def getAnalyzeOfSelect(self):
        if not QusAndAns.sheet1: return []
        if QusAndAns.analyze_select: return QusAndAns.analyze_select
        for row in range(3,3+QusAndAns.cnt_select):
            con=QusAndAns.sheet1['H%d'%row:'H%d'%row]
            QusAndAns.analyze_select.append(str(con[0][0].value))
        return QusAndAns.analyze_select

    # 获取填空题
    def getQusOfBlank(self):
        if not QusAndAns.sheet2: return []
        if QusAndAns.qus_blank: return QusAndAns.qus_blank
        for row in range(3, 3 + QusAndAns.cnt_blank):
            con = QusAndAns.sheet2['A%d' % row:'B%d' % row]  # 每一行题目
            temp1 = '、' + str(con[0][1].value) + '\n'
            QusAndAns.qus_blank.append(temp1)
        return QusAndAns.qus_blank

    # 获取填空题答案
    def getAnsOfBlank(self):
        if not QusAndAns.sheet2: return []
        if QusAndAns.ans_blank: return QusAndAns.ans_blank
        for row in range(3, 3 + QusAndAns.cnt_blank):
            con = QusAndAns.sheet2['C%d' % row:'C%d' % row]
            QusAndAns.ans_blank.append(str(con[0][0].value))
        return QusAndAns.ans_blank

    # 获取填空题解析
    def getAnalyzeOfblank(self):
        if not QusAndAns.sheet2: return []
        if QusAndAns.analyze_blank: return QusAndAns.analyze_blank
        for row in range(3, 3 + QusAndAns.cnt_blank):
            con = QusAndAns.sheet2['D%d' % row:'D%d' % row]
            QusAndAns.analyze_blank.append(str(con[0][0].value))
        return QusAndAns.analyze_blank

    # 新增：获取判断题
    def getQusOfJudge(self):
        if not QusAndAns.sheet4: return []
        if QusAndAns.qus_judge: return QusAndAns.qus_judge
        for row in range(3, 3 + QusAndAns.cnt_judge):
            con = QusAndAns.sheet4['A%d' % row:'B%d' % row]
            temp1 = '、' + str(con[0][1].value) + '\n'
            QusAndAns.qus_judge.append(temp1)
        return QusAndAns.qus_judge

    # 新增：获取判断题答案
    def getAnsOfJudge(self):
        if not QusAndAns.sheet4: return []
        if QusAndAns.ans_judge: return QusAndAns.ans_judge
        for row in range(3, 3 + QusAndAns.cnt_judge):
            con = QusAndAns.sheet4['C%d' % row:'C%d' % row]
            # 答案转换为 'T' 或 'F'
            ans_val = str(con[0][0].value).upper()
            if '对' in ans_val or 'T' in ans_val:
                QusAndAns.ans_judge.append('T')
            else:
                QusAndAns.ans_judge.append('F')
        return QusAndAns.ans_judge

    # 新增：获取判断题解析
    def getAnalyzeOfJudge(self):
        if not QusAndAns.sheet4: return []
        if QusAndAns.analyze_judge: return QusAndAns.analyze_judge
        for row in range(3, 3 + QusAndAns.cnt_judge):
            con = QusAndAns.sheet4['D%d' % row:'D%d' % row]
            QusAndAns.analyze_judge.append(str(con[0][0].value))
        return QusAndAns.analyze_judge


    # 生成选择题题目的随机顺序
    def getRandQusOfSelect(self):
        if not QusAndAns.sheet1: return []
        randselect=[n for n in range(QusAndAns.cnt_select)]
        random.shuffle(randselect)
        return randselect

    # 生成选择题选项的随机顺序
    def getRandAnsOfSelect(self):
        randselect = [n for n in range(4)]
        random.shuffle(randselect)
        return randselect

    # 生成填空题题目的随机顺序
    def getRandQusOfBlank(self):
        if not QusAndAns.sheet2: return []
        randselect = [n for n in range(QusAndAns.cnt_blank)]
        random.shuffle(randselect)
        return randselect

    # 新增：生成判断题题目的随机顺序
    def getRandQusOfJudge(self):
        if not QusAndAns.sheet4: return []
        randselect = [n for n in range(QusAndAns.cnt_judge)]
        random.shuffle(randselect)
        return randselect
